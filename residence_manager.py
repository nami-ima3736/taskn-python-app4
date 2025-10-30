"""
在留資格管理システム (Residence Status Management System)
.xlsxファイルから在留資格情報を読み込み、期限管理を行うプログラム
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
import warnings
import os
warnings.filterwarnings('ignore')


DIGIT_TRANSLATION = str.maketrans('０１２３４５６７８９', '0123456789')


class ResidenceStatusManager:
    """在留資格管理クラス"""
    
    def __init__(self, excel_file_path):
        """
        初期化
        
        Args:
            excel_file_path (str): Excelファイルのパス
        """
        self.excel_file_path = excel_file_path
        self.df = None
        self.workbook = None
        self.worksheet = None
        self.last_error = None
        
    def load_excel(self):
        """Excelファイルを読み込む"""
        try:
            print(f"[DEBUG] ファイルパス: {self.excel_file_path}")
            
            # pandasでデータ読み込み（数式の計算結果を読み込む）
            print("[DEBUG] pandasでExcelを読み込み中...")
            self.df = pd.read_excel(self.excel_file_path)
            self.last_error = None
            print(f"[DEBUG] pandas読み込み成功: {len(self.df)}行, {len(self.df.columns)}列")
            print(f"[DEBUG] 列名: {list(self.df.columns)}")
            
            # openpyxlでワークブック読み込み（数式も読み込む）
            print("[DEBUG] openpyxlでExcelを読み込み中...")
            self.workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)
            self.worksheet = self.workbook.active
            print("[DEBUG] openpyxl読み込み成功")
            
            print(f"[OK] Excelファイルを読み込みました: {self.excel_file_path}")
            print(f"  データ件数: {len(self.df)}件\n")
            
            return True
        except FileNotFoundError:
            print(f"[ERROR] エラー: ファイルが見つかりません: {self.excel_file_path}")
            return False
        except Exception as e:
            import traceback
            print(f"[ERROR] エラー: ファイル読み込み中にエラーが発生しました: {e}")
            print(traceback.format_exc())
            return False
    
    def calculate_days_until_expiration(self, expiration_date):
        """
        満了日までの日数を計算
        Excelの数式: =満了年月日 - TODAY()
        
        Args:
            expiration_date: 満了年月日
            
        Returns:
            int: 満了までの日数（負の値は期限切れ）
        """
        if pd.isna(expiration_date):
            return None
        
        try:
            if isinstance(expiration_date, str):
                expiration_date = pd.to_datetime(expiration_date)
            
            today = datetime.now().date()
            if isinstance(expiration_date, pd.Timestamp):
                expiration_date = expiration_date.date()
            elif isinstance(expiration_date, datetime):
                expiration_date = expiration_date.date()
            
            delta = (expiration_date - today).days
            return delta
        except Exception as e:
            print(f"日数計算エラー: {e}")
            return None
    
    def calculate_deadline_date(self, expiration_date, days_before):
        """
        期限日を計算
        Excelの数式: =満了年月日 - 設定期限
        
        Args:
            expiration_date: 満了年月日
            days_before (int): 何日前に設定するか
            
        Returns:
            date: 期限日
        """
        if pd.isna(expiration_date) or pd.isna(days_before):
            return None
        
        try:
            if isinstance(expiration_date, str):
                expiration_date = pd.to_datetime(expiration_date)
            
            if isinstance(expiration_date, pd.Timestamp):
                expiration_date = expiration_date.date()
            elif isinstance(expiration_date, datetime):
                expiration_date = expiration_date.date()
            
            deadline = expiration_date - timedelta(days=int(days_before))
            return deadline
        except Exception as e:
            print(f"期限日計算エラー: {e}")
            return None
    
    def process_data(self):
        """データを処理し、計算フィールドを追加"""
        try:
            if self.df is None:
                print("[ERROR] エラー: データが読み込まれていません")
                return False
            self.last_error = None
            
            print(f"[DEBUG] process_data開始: {len(self.df)}行")
            print(f"[DEBUG] 列名: {list(self.df.columns)}")
            
            # 満了年月日の列名を確認
            expiration_col = None
            for col in ['満了年月日', '満了日', 'expiration_date']:
                if col in self.df.columns:
                    expiration_col = col
                    break
            
            if expiration_col is None:
                print("[ERROR] エラー: '満了年月日'列が見つかりません")
                print(f"[DEBUG] 利用可能な列: {list(self.df.columns)}")
                return False
            
            print(f"[DEBUG] 満了年月日列を検出: {expiration_col}")
            
            # 指定列の数値を半角に正規化
            def normalize_digits(value):
                if pd.isna(value):
                    return value
                if isinstance(value, str):
                    return value.translate(DIGIT_TRANSLATION)
                return value

            if '期生' in self.df.columns:
                self.df['期生'] = self.df['期生'].apply(normalize_digits)
            if '在留資格' in self.df.columns:
                self.df['在留資格'] = self.df['在留資格'].apply(normalize_digits)
            if '在留カード番号' in self.df.columns:
                self.df['在留カード番号'] = self.df['在留カード番号'].apply(normalize_digits)
            if '特技1号在留期限' in self.df.columns:
                self.df['特技1号在留期限'] = self.df['特技1号在留期限'].apply(normalize_digits)

            # 特技1号在留期限列の検証
            if '特技1号在留期限' not in self.df.columns:
                msg = "Excelに『特技1号在留期限』列(S列)が存在しません"
                self.last_error = msg
                print(f"[ERROR] {msg}")
                return False

            for idx, row in self.df.iterrows():
                zairyu_shikaku = str(row.get('在留資格', '')).translate(DIGIT_TRANSLATION)
                if '特定技能' in zairyu_shikaku and '1号' in zairyu_shikaku:
                    limit_raw = row.get('特技1号在留期限')
                    if pd.isna(limit_raw) or str(limit_raw).strip() == '':
                        msg = f"特定技能1号の行で『特技1号在留期限』が未入力です (Excel行: {idx + 2})"
                        self.last_error = msg
                        print(f"[ERROR] {msg}")
                        return False
                    limit_str = str(limit_raw).strip()
                    if isinstance(limit_raw, str):
                        limit_str = limit_str.translate(DIGIT_TRANSLATION)
                    try:
                        numeric_limit = float(limit_str)
                    except ValueError:
                        msg = f"特定技能1号の行で『特技1号在留期限』に数値を入力してください (Excel行: {idx + 2})"
                        self.last_error = msg
                        print(f"[ERROR] {msg}")
                        return False
                    self.df.at[idx, '特技1号在留期限'] = numeric_limit

            # 満了日数を計算（特定技能1号の場合は既満了日数を考慮）
            print("[DEBUG] 満了日数列を計算中...")
            def calculate_manryo_days(row):
                zairyu_shikaku = str(row.get('在留資格', ''))
                # 全角数字を半角に変換
                zairyu_shikaku = zairyu_shikaku.translate(DIGIT_TRANSLATION)
                ki_manryo = row.get('既満了日数', None)
                kyoka_date = row.get('許可年月日')
                manryo_date = row.get(expiration_col)

                is_skill1 = ('特定技能' in zairyu_shikaku) and ('1号' in zairyu_shikaku)
                is_skill2 = ('特定技能' in zairyu_shikaku) and ('2号' in zairyu_shikaku)
                is_gino = zairyu_shikaku.startswith('技能実習')

                # 日付正規化
                if not pd.isna(kyoka_date):
                    if isinstance(kyoka_date, str):
                        kyoka_date = pd.to_datetime(kyoka_date).date()
                    elif isinstance(kyoka_date, pd.Timestamp):
                        kyoka_date = kyoka_date.date()
                if not pd.isna(manryo_date):
                    if isinstance(manryo_date, str):
                        manryo_date = pd.to_datetime(manryo_date).date()
                    elif isinstance(manryo_date, pd.Timestamp):
                        manryo_date = manryo_date.date()

                if pd.isna(kyoka_date) or pd.isna(manryo_date):
                    return None

                base = (manryo_date - kyoka_date).days + 1

                if is_gino or is_skill2:
                    # 技能実習* または 特定技能2号は満了日数を空白（None）にする
                    return None
                elif is_skill1:
                    # 特定技能1号は既満了日数が必須（入力側で検証）。欠落時は0として扱う。
                    if pd.isna(ki_manryo) or ki_manryo == '':
                        ki_manryo = 0
                    return int(ki_manryo) + base
                else:
                    # その他: 既満了日数が数値なら加算、なければ空白
                    if pd.isna(ki_manryo) or ki_manryo == '':
                        return None
                    return int(ki_manryo) + base
            
            self.df['満了日数'] = self.df.apply(calculate_manryo_days, axis=1)
            
            # 期限日1-3を常に再計算（設定期限に基づいて最新の値を計算）
            for i in range(1, 4):
                setting_col = f'設定期限{i}'
                deadline_col = f'期限日{i}'
                
                if setting_col in self.df.columns:
                    print(f"[DEBUG] {deadline_col}列を計算中...")
                    self.df[deadline_col] = self.df.apply(
                        lambda row: self.calculate_deadline_date(
                            row[expiration_col], 
                            row[setting_col]
                        ) if setting_col in row else None,
                        axis=1
                    )
            
            print("[OK] データ処理が完了しました\n")
            return True
        except Exception as e:
            import traceback
            print(f"[ERROR] データ処理エラー: {e}")
            print(traceback.format_exc())
            return False
    
    def get_expiring_soon(self, days_threshold=30):
        """
        期限日を超過しているデータを取得
        満了年月日までの残り日数が、期限日1/2/3のいずれかを超過している場合にアラート対象
        
        Args:
            days_threshold (int): 使用しない（互換性のため残す）
            
        Returns:
            DataFrame: アラート対象のデータ
        """
        if self.df is None or '満了年月日' not in self.df.columns:
            return pd.DataFrame()
        
        today = datetime.now().date()
        alert_indices = []
        
        for idx, row in self.df.iterrows():
            # 満了年月日までの残り日数を計算
            expiration_date = row.get('満了年月日')
            if pd.isna(expiration_date):
                continue
            
            if isinstance(expiration_date, (datetime, pd.Timestamp)):
                exp_date = expiration_date.date()
            else:
                exp_date = pd.to_datetime(expiration_date).date()
            
            days_to_expiration = (exp_date - today).days
            
            # 期限日1/2/3のいずれかを超過しているかチェック
            for i in range(1, 4):
                setting_col = f'設定期限{i}'
                if setting_col in row and not pd.isna(row[setting_col]):
                    threshold = row[setting_col]
                    if days_to_expiration <= threshold:
                        alert_indices.append(idx)
                        break
        
        if not alert_indices:
            return pd.DataFrame()
        
        # アラート対象のデータを抽出
        expiring = self.df.loc[alert_indices].copy()
        
        # 満了年月日までの残り日数を追加
        expiring['満了年月日までの残り日数'] = expiring['満了年月日'].apply(
            lambda x: (pd.to_datetime(x).date() - today).days if not pd.isna(x) else None
        )
        
        # 残り日数でソート（期限が近い順）
        expiring = expiring.sort_values('満了年月日までの残り日数')
        
        return expiring
    
    def display_summary(self):
        """データのサマリーを表示"""
        if self.df is None:
            print("[ERROR] データがありません")
            return
        
        print("=" * 80)
        print("在留資格管理サマリー")
        print("=" * 80)
        print(f"総データ件数: {len(self.df)}件")
        
        if '満了日数' in self.df.columns:
            valid_data = self.df[self.df['満了日数'].notna()]
            expired = valid_data[valid_data['満了日数'] < 0]
            expiring_30 = valid_data[(valid_data['満了日数'] >= 0) & (valid_data['満了日数'] <= 30)]
            expiring_60 = valid_data[(valid_data['満了日数'] > 30) & (valid_data['満了日数'] <= 60)]
            expiring_90 = valid_data[(valid_data['満了日数'] > 60) & (valid_data['満了日数'] <= 90)]
            
            print(f"\n【期限状況】")
            print(f"  [!] 期限切れ: {len(expired)}件")
            print(f"  [警告] 30日以内: {len(expiring_30)}件")
            print(f"  [警告] 31-60日以内: {len(expiring_60)}件")
            print(f"  [注意] 61-90日以内: {len(expiring_90)}件")
        
        print("=" * 80)
        print()
    
    def display_expiring_list(self, days_threshold=90):
        """
        期限が近いデータの一覧を表示
        
        Args:
            days_threshold (int): 何日以内のデータを表示するか
        """
        expiring = self.get_expiring_soon(days_threshold)
        
        if len(expiring) == 0:
            print(f"[OK] {days_threshold}日以内に期限を迎えるデータはありません\n")
            return
        
        print(f"\n{'=' * 80}")
        print(f"{days_threshold}日以内に期限を迎えるデータ ({len(expiring)}件)")
        print(f"{'=' * 80}\n")
        
        # 表示する列を選択
        display_cols = []
        for col in ['担当者コード', '氏名１', '氏名２', '在留資格', '満了年月日', '満了日数', 
                    '期限日1', '期限日2', '期限日3']:
            if col in expiring.columns:
                display_cols.append(col)
        
        for idx, row in expiring.iterrows():
            days = row['満了日数']
            
            # 期限状態によってマーカーを変更
            if days < 0:
                marker = "[!] 期限切れ"
            elif days <= 7:
                marker = "[緊急] 緊急"
            elif days <= 30:
                marker = "[警告] 警告"
            else:
                marker = "[注意] 注意"
            
            print(f"{marker}")
            for col in display_cols:
                value = row[col]
                if pd.isna(value):
                    value = "-"
                elif isinstance(value, (datetime, pd.Timestamp)):
                    value = value.strftime('%Y-%m-%d')
                print(f"  {col}: {value}")
            print()
    
    def save_processed_data(self, output_path=None):
        """
        処理済みデータを保存（計算式を含む）
        
        Args:
            output_path (str): 出力ファイルパス（Noneの場合は元のファイルを上書き）
        """
        if self.df is None:
            print("[ERROR] 保存するデータがありません")
            return False
        
        if output_path is None:
            output_path = self.excel_file_path.replace('.xlsx', '_processed.xlsx')
        
        # 既存のファイルが開かれている場合に備えて削除を試みる
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
                print(f"[DEBUG] 既存ファイルを削除: {output_path}")
            except Exception as e:
                print(f"[WARNING] 既存ファイルの削除に失敗（ファイルが開かれている可能性）: {e}")
                # 別名で保存を試みる
                import time
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                output_path = output_path.replace('.xlsx', f'_{timestamp}.xlsx')
                print(f"[INFO] 別名で保存します: {output_path}")
        
        try:
            # 日付列から時間情報を削除（日付のみにする）
            df_to_save = self.df.copy()
            date_columns = ['許可年月日', '満了年月日', '期限日1', '期限日2', '期限日3', '生年月日']
            for col in date_columns:
                if col in df_to_save.columns:
                    df_to_save[col] = df_to_save[col].apply(
                        lambda x: x.date() if isinstance(x, (datetime, pd.Timestamp)) and not pd.isna(x) else x
                    )
            
            # 担当者コードはそのまま保持（変換しない）
            
            # 期生に「期」を付加（まだ付いていない場合）
            if '期生' in df_to_save.columns:
                df_to_save['期生'] = df_to_save['期生'].apply(
                    lambda x: f"{x}期" if pd.notna(x) and str(x).strip() != '' and not str(x).endswith('期') else x
                )
            
            # 既満了日数が0の場合は空白（None）に変換
            # 在留資格によって出力仕様を変更
            if '既満了日数' in df_to_save.columns and '在留資格' in df_to_save.columns:
                def _format_ki(row):
                    z = str(row.get('在留資格', ''))
                    z = z.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
                    x = row.get('既満了日数')
                    is_skill1 = ('特定技能' in z) and ('1号' in z)
                    is_skill2 = ('特定技能' in z) and ('2号' in z)
                    is_gino = z.startswith('技能実習')
                    
                    # 技能実習* または 特定技能2号 の場合は常に空白
                    if is_gino or is_skill2:
                        return None
                        
                    # 特定技能1号の場合は0以上の数値をそのまま出力（空白にしない）
                    if is_skill1:
                        return 0 if (pd.isna(x) or x == '') else x
                        
                    # その他の在留資格は0の場合は空白、それ以外はそのまま
                    return (None if (pd.notna(x) and x == 0) else x)
                    
                df_to_save['既満了日数'] = df_to_save.apply(_format_ki, axis=1)

            # 担当者コードは数値として出力
            if '担当者コード' in df_to_save.columns:
                df_to_save['担当者コード'] = pd.to_numeric(df_to_save['担当者コード'], errors='coerce')
            
            # まずpandasでExcelファイルを作成
            df_to_save.to_excel(output_path, index=False)
            
            # openpyxlでファイルを開いて計算式を設定
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # 列名のインデックスを取得
            header_row = [cell.value for cell in ws[1]]
            
            # 日付列のフォーマットを設定 (yyyy/mm/dd 形式、2桁の月・日でゼロパディング)
            date_style = openpyxl.styles.NamedStyle(name='custom_date_style', number_format='yyyy/mm/dd')
            
            # 日付列のインデックスを取得
            date_columns = []
            for col_idx, col_name in enumerate(header_row, 1):
                if col_name and any(date_col in str(col_name) for date_col in ['年月日', '期日', '生年月日']):
                    date_columns.append(col_idx)
            
            # 日付列にスタイルを適用
            for col_idx in date_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:  # ヘッダー行を除く
                    if cell.value and isinstance(cell.value, (datetime, datetime.date)):
                        cell.style = date_style
            
            # 満了日数列のインデックスを取得
            manryo_days_col = None
            if '満了日数' in header_row:
                manryo_days_col = header_row.index('満了日数') + 1
            
            # 期限日列のインデックスを取得
            deadline_cols = {}
            for i in range(1, 4):
                col_name = f'期限日{i}'
                if col_name in header_row:
                    deadline_cols[i] = header_row.index(col_name) + 1
            
            # 各行をチェックして計算式を設定
            for row_idx in range(2, ws.max_row + 1):
                # 在留資格を取得
                zairyu_shikaku_col = header_row.index('在留資格') + 1 if '在留資格' in header_row else None
                zairyu_shikaku = ws.cell(row=row_idx, column=zairyu_shikaku_col).value if zairyu_shikaku_col else None
                
                # 満了日数に計算式を設定（すべての行に設定）
                if manryo_days_col:
                    # 列参照を取得
                    col_map = {}
                    for col_name in ['既満了日数', '許可年月日', '満了年月日', '在留資格']:
                        if col_name in header_row:
                            col_idx = header_row.index(col_name) + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            col_map[col_name] = f'{col_letter}{row_idx}'
                    
                    # 計算式を設定
                    # カテゴリ別の計算式を設定
                    if all(k in col_map for k in ['既満了日数', '許可年月日', '満了年月日', '在留資格']):
                        ki_cell = col_map["既満了日数"]
                        manryo_cell = col_map["満了年月日"]
                        kyoka_cell = col_map["許可年月日"]
                        zairyu_cell = col_map["在留資格"]
                        # 条件式
                        cond_dates = f"OR({manryo_cell}=\"\",{kyoka_cell}=\"\")"
                        cond_gino = f"LEFT({zairyu_cell},4)=\"技能実習\""
                        cond_skill1 = f"AND(ISNUMBER(SEARCH(\"特定技能\",{zairyu_cell})),OR(ISNUMBER(SEARCH(\"1号\",{zairyu_cell})),ISNUMBER(SEARCH(\"１号\",{zairyu_cell}))))"
                        cond_skill2 = f"AND(ISNUMBER(SEARCH(\"特定技能\",{zairyu_cell})),OR(ISNUMBER(SEARCH(\"2号\",{zairyu_cell})),ISNUMBER(SEARCH(\"２号\",{zairyu_cell}))))"
                        
                        # 計算式の基本部分（全ケースで同じ）
                        base_formula = f"{ki_cell}+({manryo_cell}-{kyoka_cell}+1)"
                        
                        # 各条件に応じた計算式
                        # 1. 日付が空の場合は空白
                        # 2. 技能実習* または 特定技能2号 の場合: 空白を返す
                        # 3. 特定技能1号 または その他: 既満了日数 + (満了-許可+1)
                        formula = (
                            f"=IF({cond_dates},\"\","
                            f"IF(OR({cond_gino},{cond_skill2}),"
                            f"\"\","  # 技能実習* または 特定技能2号は空白
                            f"IF({ki_cell}=\"\",\"\",{base_formula})))"  # その他（特定技能1号含む）
                        )
                        ws.cell(row=row_idx, column=manryo_days_col).value = formula
                
                # 期限日1/2/3に計算式を設定
                manryo_col = header_row.index('満了年月日') + 1 if '満了年月日' in header_row else None
                if manryo_col:
                    manryo_letter = openpyxl.utils.get_column_letter(manryo_col)
                    
                    for i, col_idx in deadline_cols.items():
                        setting_col_name = f'設定期限{i}'
                        if setting_col_name in header_row:
                            setting_col_idx = header_row.index(setting_col_name) + 1
                            setting_letter = openpyxl.utils.get_column_letter(setting_col_idx)
                            # 計算式: =満了年月日-設定期限
                            formula = f"={manryo_letter}{row_idx}-{setting_letter}{row_idx}"
                            ws.cell(row=row_idx, column=col_idx).value = formula
            
            # ファイルを保存
            wb.save(output_path)
            print(f"[OK] 処理済みデータを保存しました: {output_path}\n")
            
            # 日付列のフォーマットを直接修正する（openpyxlのスタイルが効かない場合のフォールバック）
            try:
                from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900
                from openpyxl.styles import numbers
                
                wb = openpyxl.load_workbook(output_path)
                ws = wb.active
                
                # 日付列を特定
                date_cols = []
                for idx, cell in enumerate(ws[1], 1):  # ヘッダー行をチェック
                    if any(date_word in str(cell.value) for date_word in ['年月日', '期日', '生年月日']):
                        date_cols.append(idx)
                
                # 日付フォーマットを適用
                date_format = 'yyyy/mm/dd;@'
                for col_idx in date_cols:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[col_letter]:
                        if cell.row == 1:  # ヘッダー行はスキップ
                            continue
                        if cell.value and isinstance(cell.value, (datetime, datetime.date)):
                            cell.number_format = date_format
                
                # 期限日1,2,3列を明示的にyyyy/mm/dd形式でフォーマット
                for i in range(1, 4):
                    deadline_col_name = f'期限日{i}'
                    if deadline_col_name in header_row:
                        col_idx = header_row.index(deadline_col_name) + 1
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        for cell in ws[col_letter]:
                            if cell.row == 1:  # ヘッダー行はスキップ
                                continue
                            cell.number_format = date_format
                
                wb.save(output_path)
                print(f"[INFO] 日付フォーマットを適用しました: {output_path}\n")
                
            except Exception as e:
                print(f"[WARNING] 日付フォーマットの適用中にエラーが発生しました: {e}")
            
            return True
        except Exception as e:
            print(f"[ERROR] 保存エラー: {e}")
            return False
    
    def export_alert_list(self, output_path, days_threshold=30):
        """
        アラート対象のデータをExcelファイルとして出力
        
        Args:
            output_path (str): 出力ファイルパス
            days_threshold (int): 何日以内のデータを出力するか
        """
        expiring = self.get_expiring_soon(days_threshold)
        
        if len(expiring) == 0:
            print(f"[OK] 期限日を超過しているデータはありません")
            return False
        
        try:
            # 日付列から時間情報を削除（日付のみにする）
            date_columns = ['許可年月日', '満了年月日', '期限日1', '期限日2', '期限日3', '生年月日']
            if '満了年月日までの残り日数' in expiring.columns:
                expiring = expiring.drop(columns=['満了年月日までの残り日数'])
            for col in date_columns:
                if col in expiring.columns:
                    expiring[col] = expiring[col].apply(
                        lambda x: x.date() if isinstance(x, (datetime, pd.Timestamp)) and not pd.isna(x) else x
                    )
            
            expiring.to_excel(output_path, index=False)
            print(f"[OK] アラートリストを出力しました: {output_path}")
            print(f"  対象件数: {len(expiring)}件\n")
            return True
        except Exception as e:
            print(f"[ERROR] 出力エラー: {e}")
            return False


def main():
    """メイン処理"""
    print("\n" + "=" * 80)
    print("在留資格管理システム")
    print("=" * 80 + "\n")
    
    # Excelファイルのパスを指定
    excel_file = "在留資格管理.xlsx"
    
    # 管理システムのインスタンス作成
    manager = ResidenceStatusManager(excel_file)
    
    # データ読み込み
    if not manager.load_excel():
        return
    
    # データ処理
    if not manager.process_data():
        return
    
    # サマリー表示
    manager.display_summary()
    
    # 期限が近いデータを表示（90日以内）
    manager.display_expiring_list(days_threshold=90)
    
    # 処理済みデータを保存
    manager.save_processed_data()
    
    # アラートリストを出力（30日以内）
    manager.export_alert_list("アラートリスト.xlsx", days_threshold=30)
    
    print("=" * 80)
    print("処理が完了しました")
    print("=" * 80 + "\n")


if __name__ == "__main__":
    main()
